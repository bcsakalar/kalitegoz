"""SMTP zamanlanmis e-posta raporu (Dalga 7).

Haftalik ekip performans raporunu (Excel) uretir ve SMTP ile gonderir.
Kimlik bilgisi (SMTP_HOST vb.) tanimli degilse GONDERMEZ — raporu yine uretir
ve log'a "gonderilmedi (SMTP yapilandirilmamis)" yazar. Boylece ozellik
kimlik bilgisi olmadan da guvenle calisir ve test edilebilir.
"""

from __future__ import annotations

import io
import logging
import smtplib
from datetime import datetime
from email.message import EmailMessage

from sqlalchemy import Integer, cast, func
from sqlalchemy.orm import Session

from ..config import settings
from ..models import Agent, Call, CallStatus, Team, Tenant

logger = logging.getLogger(__name__)


def build_team_xlsx(db: Session, tenant_id: int) -> bytes:
    """Ekip performans raporunu Excel bayti olarak uretir.

    reports.py'daki endpoint ile ayni veri; burada dosya olarak degil bellek
    bayti olarak dondurulur ki e-postaya eklenebilsin.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    done = (Call.agent_id == Agent.id) & (Call.status == CallStatus.done)
    rows = (
        db.query(
            Agent.name, Agent.team_id,
            func.count(Call.id), func.avg(Call.total_score),
            func.avg(Call.predicted_csat),
            func.sum(cast(Call.zeroed, Integer)), func.sum(cast(Call.is_crisis, Integer)),
        )
        .outerjoin(Call, done)
        .filter(Agent.tenant_id == tenant_id)
        .group_by(Agent.name, Agent.team_id)
        .order_by(func.avg(Call.total_score).desc().nulls_last())
        .all()
    )
    team_names = {t.id: t.name for t in db.query(Team).filter(Team.tenant_id == tenant_id).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Ekip Raporu"
    ws.append(["Temsilci", "Takım", "Çağrı", "Ort. Puan", "Ort. CSAT", "Sıfırlanan", "Kriz"])
    fill = PatternFill("solid", fgColor="2A78D6")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for name, tid, cnt, avg, csat, zeroed, crisis in rows:
        ws.append([
            name, team_names.get(tid, "-"), cnt or 0,
            round(avg, 1) if avg is not None else "-",
            round(csat, 1) if csat is not None else "-",
            int(zeroed or 0), int(crisis or 0),
        ])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = width + 3

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_message(tenant_name: str, xlsx: bytes, recipients: list[str]) -> EmailMessage:
    msg = EmailMessage()
    today = datetime.now().strftime("%d.%m.%Y")
    msg["Subject"] = f"KaliteGöz — Haftalık Ekip Raporu ({tenant_name}, {today})"
    msg["From"] = settings.smtp_from
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Merhaba,\n\n{tenant_name} için haftalık ekip kalite raporu ektedir.\n\n"
        "Bu e-posta KaliteGöz tarafından otomatik gönderilmiştir.\n"
    )
    msg.add_attachment(
        xlsx, maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"ekip_raporu_{datetime.now():%Y%m%d}.xlsx",
    )
    return msg


def send_report(db: Session, tenant_id: int, tenant_name: str,
                recipients: list[str] | None = None) -> dict:
    """Bir tenant icin ekip raporunu uretir ve (yapilandirilmissa) e-postayla yollar.

    Doner: {"generated": True, "sent": bool, "recipients": [...], "reason": ...}
    SMTP veya alici yoksa 'sent': False ile doner (hata degil — opsiyonel ozellik).
    """
    recipients = recipients if recipients is not None else settings.report_recipient_list
    xlsx = build_team_xlsx(db, tenant_id)
    result = {"generated": True, "bytes": len(xlsx), "sent": False, "recipients": recipients}

    if not settings.smtp_host:
        result["reason"] = "SMTP yapilandirilmamis (SMTP_HOST bos)"
        logger.info("Rapor uretildi ama gonderilmedi: %s", result["reason"])
        return result
    if not recipients:
        result["reason"] = "Alici yok (REPORT_RECIPIENTS bos)"
        logger.info("Rapor uretildi ama gonderilmedi: %s", result["reason"])
        return result

    try:
        msg = _build_message(tenant_name, xlsx, recipients)
        with smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=30) as server:
            if settings.smtp_use_tls:
                server.starttls()
            if settings.smtp_user:
                server.login(settings.smtp_user, settings.smtp_password)
            server.send_message(msg)
        result["sent"] = True
        logger.info("Haftalik rapor gonderildi: tenant=%s alici=%s", tenant_name, recipients)
    except Exception as exc:  # SMTP hatasi raporu dusurmemeli
        result["reason"] = f"SMTP hatasi: {exc}"
        logger.warning("Rapor gonderilemedi (%s): %s", tenant_name, exc)
    return result


def send_all_tenants(db: Session) -> dict:
    """Tum aktif tenant'lar icin haftalik rapor (beat task bunu cagirir)."""
    out = {}
    for tenant in db.query(Tenant).filter(Tenant.is_active.is_(True)).all():
        out[tenant.slug] = send_report(db, tenant.id, tenant.name)
    return out
