"""Rapor merkezi: Excel (ekip raporu) ve PDF (temsilci karnesi) export."""

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import Integer, cast, func
from sqlalchemy.orm import Session

from ..config import settings
from ..db import get_db
from ..deps import CurrentUser, require_admin, require_staff
from ..models import Agent, Call, CallStatus, Score, Team

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])


@router.get("/email/status")
def email_report_status(user: CurrentUser = Depends(require_staff)):
    """E-posta raporu yapilandirmasi (arayuz durum gostergesi)."""
    return {
        "smtp_configured": bool(settings.smtp_host),
        "recipients": settings.report_recipient_list,
        "schedule": "Pazartesi 08:30",
    }


@router.post("/email/send-now")
def send_report_now(db: Session = Depends(get_db), user: CurrentUser = Depends(require_admin)):
    """Haftalik ekip raporunu simdi uret ve (yapilandirilmissa) gonder (admin)."""
    from ..services import email_reports

    tenant_name = "Tenant"
    from ..models import Tenant
    t = db.get(Tenant, user.tenant_id)
    if t:
        tenant_name = t.name
    return email_reports.send_report(db, user.tenant_id, tenant_name)


@router.get("/team.xlsx")
def team_report_xlsx(team_id: int | None = None, db: Session = Depends(get_db),
                     user: CurrentUser = Depends(require_staff)):
    """Ekip/sirket geneli temsilci performans raporu (Excel)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    done = (Call.agent_id == Agent.id) & (Call.status == CallStatus.done)
    q = (
        db.query(
            Agent.name, Agent.team_id,
            func.count(Call.id), func.avg(Call.total_score),
            func.avg(Call.predicted_csat),
            func.sum(cast(Call.zeroed, Integer)), func.sum(cast(Call.is_crisis, Integer)),
        )
        .outerjoin(Call, done)
        .filter(Agent.tenant_id == user.tenant_id)
    )
    if user.role.value == "supervisor" and user.team_id:
        team_id = user.team_id
    if team_id:
        q = q.filter(Agent.team_id == team_id)
    rows = q.group_by(Agent.name, Agent.team_id).order_by(func.avg(Call.total_score).desc().nulls_last()).all()
    team_names = {t.id: t.name for t in db.query(Team).filter(Team.tenant_id == user.tenant_id).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Ekip Raporu"
    headers = ["Temsilci", "Takım", "Çağrı", "Ort. Puan", "Ort. CSAT", "Sıfırlanan", "Kriz"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="2A78D6")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for name, tid, cnt, avg, csat, zeroed, crisis in rows:
        ws.append([
            name, team_names.get(tid, "-"), cnt or 0,
            round(avg, 1) if avg is not None else "-",
            round(csat, 1) if csat is not None else "-",
            int(zeroed or 0), int(crisis or 0),
        ])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = width + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ekip_raporu_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/agent/{agent_id}.pdf")
def agent_scorecard_pdf(agent_id: int, db: Session = Depends(get_db),
                        user: CurrentUser = Depends(require_staff)):
    """Temsilci karnesi (PDF)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    agent = db.query(Agent).filter(
        Agent.id == agent_id, Agent.tenant_id == user.tenant_id).first()
    if agent is None:
        raise HTTPException(404, "Temsilci bulunamadi")

    done = (Call.agent_id == agent_id) & (Call.status == CallStatus.done)
    stats = db.query(
        func.count(Call.id), func.avg(Call.total_score), func.avg(Call.predicted_csat),
        func.sum(cast(Call.zeroed, Integer)), func.sum(cast(Call.is_crisis, Integer)),
    ).filter(done).one()
    crit_rows = (
        db.query(Score.criterion_name, func.avg(Score.score))
        .join(Call, Score.call_id == Call.id).filter(done)
        .group_by(Score.criterion_name).order_by(func.avg(Score.score)).all()
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"Karne - {agent.name}")
    styles = getSampleStyleSheet()
    story = [
        Paragraph("KaliteGöz — Temsilci Karnesi", styles["Title"]),
        Paragraph(f"Temsilci: <b>{agent.name}</b>", styles["Normal"]),
        Paragraph(f"Rapor tarihi: {datetime.now():%d.%m.%Y %H:%M}", styles["Normal"]),
        Spacer(1, 0.6 * cm),
    ]
    summary = [
        ["Değerlendirilen çağrı", str(stats[0] or 0)],
        ["Ortalama puan", f"{round(stats[1], 1) if stats[1] is not None else '-'}"],
        ["Ortalama CSAT (1-5)", f"{round(stats[2], 1) if stats[2] is not None else '-'}"],
        ["Sıfırlanan çağrı", str(int(stats[3] or 0))],
        ["Kriz çağrısı", str(int(stats[4] or 0))],
    ]
    t = Table(summary, colWidths=[7 * cm, 6 * cm])
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF3FB")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story += [t, Spacer(1, 0.6 * cm), Paragraph("Kriter Bazında Ortalama", styles["Heading2"])]

    crit_data = [["Kriter", "Ortalama (0-10)"]] + [
        [name, f"{round(avg, 1)}"] for name, avg in crit_rows
    ]
    ct = Table(crit_data, colWidths=[9 * cm, 4 * cm])
    ct.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2A78D6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(ct)

    doc.build(story)
    buf.seek(0)
    fname = f"karne_{agent.name}_{datetime.now():%Y%m%d}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
